from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border
import os


class Excel:

    def __init__(self, test_data_file_path):
        # 文件格式校验
        if not os.path.exists(test_data_file_path):
            print("Excel工具类初始化失败：【{}】文件不存在！".format(test_data_file_path))
            return
        if not test_data_file_path.endswith(".xlsx") or not test_data_file_path.endswith(".xlsx"):
            print("Excel工具类初始化失败：【{}】文件非excel文件类型！".format(test_data_file_path))
            return
        # 打开指定excel文件
        self.wb = load_workbook(test_data_file_path)
        # 初始化默认sheet
        self.ws = self.wb.active
        # 保存文件时使用的文件路径
        self.test_data_file_path = test_data_file_path
        # 初始化红、绿色，供样式使用
        self.color_dict = {"red": "FFFF3030", "green": "FF008B00"}

    # 查看所有sheet名称
    def get_sheets(self):
        return self.wb.sheetnames

    # 根据sheet名称切换sheet
    def change_sheet(self, sheet_name):
        if sheet_name not in self.get_sheets():
            print("sheet切换失败：【{}】指定sheet名称不存在！".format(sheet_name))
            return
        self.ws = self.wb.get_sheet_by_name(sheet_name)

    # 返回当前sheet的最大行号
    def max_row_num(self):
        return self.ws.max_row

    # 返回当前sheet的最大列号
    def max_col_num(self):
        return self.ws.max_column

    # 获取指定行数据（设定索引从0开始）
    def get_one_row_data(self, row_no):
        if row_no < 0 or row_no > self.max_row_num()-1:
            print("输入的行号【{}】有误：需在0至最大行数之间！".format(row_no))
            return
        # API的索引从1开始
        return [cell.value for cell in self.ws[row_no+1]]

    # 获取指定列数据
    def get_one_col_data(self, col_no):
        if col_no < 0 or col_no > self.max_col_num()-1:
            print("输入的列号【{}】有误：需在0至最大列数之间！".format(col_no))
            return
        return [cell.value for cell in tuple(self.ws.columns)[col_no+1]]

    # 获取当前sheet的所有行数据
    def get_all_row_data(self):
        result = []
        # # API的索引从1开始
        for row_data in self.ws[1:self.max_row_num()]:
            result.append([cell.value for cell in row_data])
        return result

    # 追加一行数据
    def write_row_data(self, data, fill_color=None, font_color=None, border=True):
        if not isinstance(data, (list, tuple)):
            print("追加的数据类型有误：需为列号或元组类型！【{}】".format(data))
            return
        self.ws.append(data)
        # 添加字体颜色
        if font_color:
            if font_color in self.color_dict.keys():
                font_color = self.color_dict[font_color]
            # 需要设置的单元格长度应与数据长度一致，否则默认与之前行的长度一致
        count = 0
        for cell in self.ws[self.max_row_num()]:
            if count > len(data) - 1:
                break
            # cell不为None，才能设置样式
            if cell:
                if cell.value in ["pass", "成功"]:
                    cell.font = Font(color=self.color_dict["green"])
                elif cell.value in ["fail", "失败"]:
                    cell.font = Font(color=self.color_dict["red"])
                else:
                    cell.font = Font(color=font_color)
            count += 1
        # 添加背景颜色
        if fill_color:
            if fill_color in self.color_dict.keys():
                fill_color = self.color_dict[fill_color]
            count = 0
            for cell in self.ws[self.max_row_num()]:
                if count > len(data) - 1:
                    break
                if cell:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                count += 1
        # 添加单元格边框
        if border:
            bd = Side(style="thin", color="000000")
            count = 0
            for cell in self.ws[self.max_row_num()]:
                if count > len(data) - 1:
                    break
                if cell:
                    cell.border = Border(left=bd, right=bd, top=bd, bottom=bd)
                count += 1

    # 保存文件
    def save(self):
        self.wb.save(self.test_data_file_path)


if __name__ == "__main__":
    from conf.global_var import *
    excel = Excel(TEST_DATA_FILE_PATH)
    excel.change_sheet("登录1")
    # print(excel.get_all_row_data())
    excel.write_row_data((1,2,"嘻哈",None,"ddd"), "red", "green")
    excel.save()